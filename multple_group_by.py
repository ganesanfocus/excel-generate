from flask import Flask, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

app = Flask(__name__)

@app.route('/download_excel')
def download_excel():
    # Updated data
    data = [
        {"name": "Ganesan", "age": 30, "gender": "Male", "class": "A", "color": "white"},
        {"name": "Mathi", "age": 30, "gender": "Male", "class": "B", "color": "black"},
        {"name": "Raja", "age": 24, "gender": "Male", "class": "A", "color": "white"},
        {"name": "Sukanya", "age": 28, "gender": "Female", "class": "B", "color": "black"},
        {"name": "Lithiya", "age": 28, "gender": "Female", "class": "A", "color": "white"},
        {"name": "Ganesan", "age": 30, "gender": "Male", "class": "A", "color": "white"},
        {"name": "Mathi", "age": 30, "gender": "Male", "class": "B", "color": "block"},
        {"name": "Raja", "age": 24, "gender": "Male", "class": "A", "color": "white"},
        {"name": "Sukanya", "age": 28, "gender": "Female", "class": "B", "color": "white"},
        {"name": "Lithiya", "age": 28, "gender": "Female", "class": "A", "color": "block"}
    ]
    
    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    # Sort by gender, class, and color for proper grouping
    df.sort_values(by=['gender', 'class', 'color'], inplace=True)
    
    # Rearrange the columns to ensure correct order
    df = df[['gender', 'class', 'color', 'name', 'age']]
    
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Write the DataFrame to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
    
    # Merge cells for gender
    merge_cells(ws, 1)  # Column index for gender
    # Merge cells for class
    merge_cells(ws, 2)  # Column index for class
    # Merge cells for color
    merge_cells(ws, 3)  # Column index for color
    
    # Specify the file path
    file_path = 'grouped_data_merged.xlsx'
    
    # Save the workbook
    wb.save(file_path)
    
    # Send the file for download
    return send_file(file_path, as_attachment=True)

def merge_cells(ws, col_idx):
    current_value = None
    start_row = None
    
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        cell = row[col_idx - 1]  # Adjust index for zero-based
    
        if cell.value != current_value:
            # Merge the previous group
            if start_row is not None:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row[0].row-1, end_column=col_idx)
                ws.cell(start_row, col_idx).alignment = Alignment(vertical='center', horizontal='center')
            
            current_value = cell.value
            start_row = row[0].row
    
    # Merge the last group
    if start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)
        ws.cell(start_row, col_idx).alignment = Alignment(vertical='center', horizontal='center')

if __name__ == '__main__':
    app.run(debug=True)
