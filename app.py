from flask import Flask, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

app = Flask(__name__)

@app.route('/download_excel')
def download_excel():
    # Original data
    data = [
        {"name": "Ganesan", "age": 30, "gender": "Male", "class" : "A"},
        {"name": "Mathi", "age": 30, "gender": "Male", "class" : "B"},
        {"name": "Raja", "age": 24, "gender": "Male", "class" : "A"},
        {"name": "Sukanya", "age": 28, "gender": "Female", "class" : "B"},
        {"name": "Lithiya", "age": 28, "gender": "Female", "class" : "A"}
    ]
    
    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    # Group by gender
    grouped_df = df.groupby('gender', sort=False).apply(lambda x: x.drop('gender', axis=1)).reset_index()
    
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    
    # Write the DataFrame to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(grouped_df, index=False, header=True), 1):
        ws.append(row)
    
    # Start merging cells
    current_gender = None
    start_row = None
    
    for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
        gender_cell = row[0]
        
        if gender_cell.value != current_gender:
            # If the gender changes, merge the previous cells
            if start_row is not None:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row[0].row-1, end_column=1)
                ws.cell(start_row, 1).alignment = Alignment(vertical='center', horizontal='center')
            
            current_gender = gender_cell.value
            start_row = row[0].row
    
    # Merge the last group
    if start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
        ws.cell(start_row, 1).alignment = Alignment(vertical='center', horizontal='center')
    
    # Specify the file path
    file_path = 'grouped_data_merged.xlsx'
    
    # Save the workbook
    wb.save(file_path)
    
    # Send the file for download
    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
